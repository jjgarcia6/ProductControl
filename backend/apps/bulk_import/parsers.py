"""Parseo de archivos CSV/Excel a filas normalizadas (capability bulk-import, F7).

Funciones puras sin dependencia del ORM: convierten el archivo subido en una lista de
`(row_number, {columna: valor})` conservando el número de fila del archivo (la cabecera
es la fila 1). Validan el formato, el tamaño y el límite de filas ANTES de procesar el
contenido, emitiendo `400` por el contrato de errores uniforme (mapeado al campo `file`).
El archivo se procesa en memoria y NO se persiste (Cloud Run stateless).
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable, Sequence

from openpyxl import load_workbook
from rest_framework import serializers

from .constants import (
    CSV_EXTENSIONS,
    EXCEL_EXTENSIONS,
    MAX_FILE_SIZE_BYTES,
    MAX_ROWS,
)

# Una fila de datos: su número en el archivo y el mapa columna -> valor (texto crudo).
ParsedRow = tuple[int, dict[str, str]]


def parse_file(uploaded_file: object) -> list[ParsedRow]:
    """Convierte el archivo subido en filas normalizadas o rechaza con 400.

    `uploaded_file` es un `UploadedFile` de Django (con `.name`, `.size` y contenido
    binario). El formato se resuelve por la extensión del nombre.
    """
    name = str(getattr(uploaded_file, "name", "") or "").lower()
    size = getattr(uploaded_file, "size", None)
    if isinstance(size, int) and size > MAX_FILE_SIZE_BYTES:
        max_mib = MAX_FILE_SIZE_BYTES // (1024 * 1024)
        raise serializers.ValidationError(
            {"file": [f"El archivo supera el tamaño máximo de {max_mib} MiB."]}
        )

    if name.endswith(CSV_EXTENSIONS):
        rows = _parse_csv(uploaded_file)
    elif name.endswith(EXCEL_EXTENSIONS):
        rows = _parse_excel(uploaded_file)
    else:
        raise serializers.ValidationError(
            {"file": ["Formato no soportado. Suba un archivo CSV o Excel (.xlsx)."]}
        )

    if len(rows) > MAX_ROWS:
        raise serializers.ValidationError(
            {"file": [f"El archivo supera el límite de {MAX_ROWS} filas. Divídalo en lotes."]}
        )
    return rows


def _parse_csv(uploaded_file: object) -> list[ParsedRow]:
    raw = uploaded_file.read()  # type: ignore[attr-defined]
    text = raw.decode("utf-8-sig") if isinstance(raw, bytes) else str(raw)
    return _rows_from_iterable(csv.reader(io.StringIO(text)))


def _parse_excel(uploaded_file: object) -> list[ParsedRow]:
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception:
        raise serializers.ValidationError({"file": ["No se pudo leer el archivo Excel."]}) from None
    try:
        worksheet = workbook.active
        if worksheet is None:
            raise serializers.ValidationError({"file": ["El archivo Excel no tiene hojas."]})
        return _rows_from_iterable(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def _rows_from_iterable(iterable: Iterable[Sequence[object]]) -> list[ParsedRow]:
    """Toma la primera fila como cabecera y devuelve las de datos con su número."""
    rows: list[ParsedRow] = []
    header: list[str] | None = None
    for line_number, values in enumerate(iterable, start=1):
        cells = [("" if value is None else str(value)).strip() for value in values]
        if header is None:
            header = cells
            continue
        if all(cell == "" for cell in cells):
            continue  # se omiten las filas totalmente vacías
        record = {
            key: (cells[col] if col < len(cells) else "") for col, key in enumerate(header) if key
        }
        rows.append((line_number, record))

    if header is None:
        raise serializers.ValidationError({"file": ["El archivo está vacío."]})
    return rows
